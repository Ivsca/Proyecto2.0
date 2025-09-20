from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from .models import Ganado, TablaRazas, Cultivo, TipoCultivo, Fertilizacion
import traceback
from io import BytesIO
import openpyxl 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import re

def PlantillaTablas(request):
    razas = TablaRazas.objects.all()
    campos_ganado = [field.name for field in Ganado._meta.fields if field.name not in ['id', 'foto']]
    return render(request, 'Ganado/tablas.html', {
        'razas': razas,
        'campos_disponibles': campos_ganado
    })

def ConsultarVacunos(request):
    try:
        # Validar y obtener parámetros
        limit = int(request.GET.get('limit', 10))
        offset = int(request.GET.get('offset', 0))
        fields = request.GET.get('fields', '')
        selected_fields = [f for f in fields.split(',') if f]
        all_fields = [f.name for f in Ganado._meta.fields]
        
        # Validar campos
        valid_fields = [f for f in selected_fields if f in all_fields]
        if not valid_fields:
            return JsonResponse({'success': True, 'vacunos': [], 'total': 0})

        queryset = Ganado.objects.all()

        # Filtro por raza
        filter_raza = request.GET.get('filter_raza')
        if filter_raza:
            queryset = queryset.filter(razas=filter_raza)

        # Orden dinámico
        order_fields = []
        for key, value in request.GET.items():
            if key.startswith('sort_') and value in ['asc', 'desc']:
                field = key[5:]
                if field in valid_fields:
                    order_fields.append(f'-{field}' if value == 'desc' else field)
        if order_fields:
            queryset = queryset.order_by(*order_fields)

        total = queryset.count()
        vacunos = queryset.only(*valid_fields)[offset:offset+limit]

        data = []
        for vacuno in vacunos:
            item = {}
            for field in valid_fields:
                value = getattr(vacuno, field, '')
                # ForeignKey: mostrar id
                if hasattr(value, 'id'):
                    value = value.id
                item[field] = value
            data.append(item)

        return JsonResponse({'success': True, 'vacunos': data, 'total': total})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def ExportarExcel(request):
    try:
        filename = request.GET.get('filename', 'ganado')
        if not filename:
            return HttpResponse("El nombre del archivo es requerido", status=400)

        if re.search(r'[<>:"/\\|?*]', filename):
            return HttpResponse("El nombre del archivo contiene caracteres no válidos", status=400)

        fields = request.GET.get('fields', '')
        selected_fields = [f for f in fields.split(',') if f]
        all_fields = [f.name for f in Ganado._meta.fields]
        valid_fields = [f for f in selected_fields if f in all_fields]
        if not valid_fields:
            return HttpResponse("No hay columnas válidas", status=400)

        queryset = Ganado.objects.all()
        filter_raza = request.GET.get('filter_raza')
        if filter_raza:
            queryset = queryset.filter(razas=filter_raza)

        order_fields = []
        for key, value in request.GET.items():
            if key.startswith('sort_') and value in ['asc', 'desc']:
                field = key[5:]
                if field in valid_fields:
                    order_fields.append(f'-{field}' if value == 'desc' else field)
        if order_fields:
            queryset = queryset.order_by(*order_fields)

        vacunos = queryset.only(*valid_fields)

        wb = Workbook()
        ws = wb.active
        ws.title = "Ganado"

        # ======================
        # ESTILOS
        # ======================
        header_fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")  # Verde
        header_font = Font(bold=True, color="FFFFFF")
        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # ======================
        # TÍTULO MERGEADO EN FILA 2
        # ======================
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(valid_fields))
        title_cell = ws.cell(row=2, column=1)
        title_cell.value = "Información del Ganado Registrado"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ======================
        # ENCABEZADOS EN FILA 3
        # ======================
        for col_idx, field in enumerate(valid_fields, 1):
            cell = ws.cell(row=3, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border_style
            cell.alignment = center_alignment

        # ======================
        # DATOS EN FILA 4+
        # ======================
        for row_idx, vacuno in enumerate(vacunos, 4):  # Comienza desde fila 4
            for col_idx, field in enumerate(valid_fields, 1):
                value = getattr(vacuno, field, '')
                if hasattr(value, 'id'):
                    value = value.id
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border_style

        # ======================
        # AJUSTAR ANCHO
        # ======================
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 3

        # ======================
        # EXPORTAR
        # ======================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        return response

    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

def PlantillaGraficas(request):
    return render(request, 'Ganado/graficas.html')

#Cultivo

def PlantillaCultivos(request):
    tipos = TipoCultivo.objects.all()
    campos_cultivo = [field.name for field in Cultivo._meta.fields if field.name not in ['id', 'foto']]
    return render(request, 'Cultivo/tablas.html', {
        'tipos': tipos,
        'campos_disponibles': campos_cultivo
    })

def ConsultarCultivos(request):
    try:
        limit = int(request.GET.get('limit', 10))
        offset = int(request.GET.get('offset', 0))
        fields = request.GET.get('fields', '')
        selected_fields = [f for f in fields.split(',') if f]
        all_fields = [f.name for f in Cultivo._meta.fields]
        valid_fields = [f for f in selected_fields if f in all_fields]

        if not valid_fields:
            return JsonResponse({'success': True, 'cultivos': [], 'total': 0})

        queryset = Cultivo.objects.all()

        # Filtro por tipo
        filter_tipo = request.GET.get('filter_tipo')
        if filter_tipo:
            queryset = queryset.filter(tipo=filter_tipo)

        # Orden dinámico
        order_fields = []
        for key, value in request.GET.items():
            if key.startswith('sort_') and value in ['asc', 'desc']:
                field = key[5:]
                if field in valid_fields:
                    order_fields.append(f'-{field}' if value == 'desc' else field)
        if order_fields:
            queryset = queryset.order_by(*order_fields)

        total = queryset.count()
        cultivos = queryset.only(*valid_fields)[offset:offset+limit]

        data = []
        for cultivo in cultivos:
            item = {}
            for field in valid_fields:
                value = getattr(cultivo, field, '')
                if hasattr(value, 'id'):
                    value = value.id
                item[field] = value
            data.append(item)

        return JsonResponse({'success': True, 'cultivos': data, 'total': total})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)




def ExportarExcelCultivos(request):
    try:
        filename = request.GET.get('filename', 'cultivos')
        if not filename or re.search(r'[<>:"/\\|?*]', filename):
            return HttpResponse("Nombre de archivo inválido", status=400)

        selected_fields = request.GET.get('fields', '').split(',')
        if not selected_fields:
            return HttpResponse("Selecciona columnas", status=400)

        # Validar campos reales y extra
        model_fields = [f.name for f in Cultivo._meta.fields]
        extra_fields = ['fecha_fertilizacion', 'dosis_fertilizacion']
        valid_fields = [f for f in selected_fields if f in model_fields or f in extra_fields]

        queryset = Cultivo.objects.all()

        # Orden y filtros (opcional)
        for key, value in request.GET.items():
            if key.startswith('sort_') and value in ['asc', 'desc']:
                field = key[5:]
                if field in valid_fields:
                    queryset = queryset.order_by(f"{'-' if value == 'desc' else ''}{field}")

        cultivos = queryset

        # Crear libro
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cultivos"

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(valid_fields))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "Información de los Cultivos Registrados"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4CAF50")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col_idx, field in enumerate(valid_fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Datos
        for row_idx, cultivo in enumerate(cultivos, start=3):
            for col_idx, field in enumerate(valid_fields, 1):
                if field in extra_fields:
                    # Buscar primera fertilización
                    fert = Fertilizacion.objects.filter(cultivo=cultivo).order_by('fecha').first()
                    if field == 'fecha_fertilizacion':
                        value = fert.fecha if fert else ''
                    elif field == 'dosis_fertilizacion':
                        value = fert.dosis if fert else ''
                else:
                    value = getattr(cultivo, field, '')
                    if hasattr(value, 'nombre_tipo'):  # relaciones
                        value = value.nombre_tipo
                ws.cell(row=row_idx, column=col_idx, value=value).border = border

        # Ancho automático
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        return response

    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)
