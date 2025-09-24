# ==========================
# Librerías estándar de Python
# ==========================
import os
import re
import json
import traceback
from datetime import datetime, date, timedelta
from io import BytesIO
from multiprocessing import connection

# ==========================
# Librerías externas
# ==========================
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================
# Django
# ==========================
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.core.paginator import Paginator
from django.core.serializers import serialize
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt


from django.contrib import messages

# ==========================
# Modelos locales
# ==========================
from .models import (
    TablaRazas,
    TipoDocumentos,
    Usuarios,
    Ganado,
    TipoCultivo,
    Cultivo,
    Enfermedades,
    TablaVacunas,
    Fertilizacion,
    TipoParcela,
    codigo,
    GanadoInactivo
)

# ==========================
# Decoradores y otros módulos locales
# ==========================
from .decoradores import login_requerido
from AppInventario.Notificaciones import NotificacionesCultivos
import hashlib, random
from datetime import timedelta
from django.utils import timezone
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from .utils import send_reset_email



# Create your views here.

# region Home
def Home(request):
    return render(request,'Home.html')
# endregion

# region Logueo

def plantilla_logue(request):
    tipos_documentos = TipoDocumentos.objects.all()
    return render(request, 'Logueo/Logueo.html', {
        'tipos_documentos': tipos_documentos
    })
# region solicitudes de acceso
@login_requerido
def TablaSolicitudesUsuarios(request):
    usuarios = Usuarios.objects.filter(estado="Solicitud")
    return render(request, 'Logueo/Table.html',{
        'usuarios':usuarios
    })

@login_requerido
def SolicitudAceptada(request, id_solicitud):
    solicitud = Usuarios.objects.get(id=id_solicitud)
    solicitud.estado = "Usuario"
    solicitud.save()
    return redirect("TablaSolicitudesUsuarios")

@login_requerido
def EliminarSolicitud(request,id_solicitud):
    Solicitud = Usuarios.objects.get(id=id_solicitud)
    Solicitud.delete()
    return redirect("TablaSolicitudesUsuarios")
# endregion
@login_requerido
def TablaUsuarios(request):
    usuarios = Usuarios.objects.filter(estado="Usuario")
    return render(request, 'Logueo/Table.html',{
        'usuarios':usuarios
    })

def Cambiarrol(request, id_usuario):
    usuario = get_object_or_404(Usuarios, id=id_usuario)

    # alternar rol
    if usuario.rol.lower() == "usuario":
        usuario.rol = "Admin"
    else:
        usuario.rol = "usuario"

    usuario.save()
    messages.success(request, f"Rol de {usuario.username} cambiado a {usuario.rol}")
    return redirect("TablaUsuarios")  # Usa el name de tu listado de usuarios

# region Register
def RegisterUser(request):
    if request.method == "POST":
        nombre_usuario = request.POST['UserName'].strip()
        nombres = request.POST['Nombres'].strip()
        apellidos = request.POST['Apellidos'].strip()
        correo = request.POST['Correo'].strip().lower()
        tipo_documento_id = request.POST['TipoDocumento']
        numero_documento = request.POST['NumeroDocumento'].strip()
        clave1 = request.POST['Clave1']
        clave2 = request.POST['Clave2']
        rol = "Usuario"
        estado = "Solicitud"  # ✅ usuario queda pendiente de aprobación admin

        # 🔹 1. Validar campos vacíos
        if not all([nombre_usuario, nombres, apellidos, correo, tipo_documento_id, numero_documento, clave1, clave2]):
            return HttpResponse("Todos los campos son obligatorios", status=400)

        # 🔹 2. Validar username
        if not re.match(r'^[a-zA-Z0-9_]{4,20}$', nombre_usuario):
            return HttpResponse("El nombre de usuario debe tener 4-20 caracteres y solo usar letras, números o _", status=400)
        if Usuarios.objects.filter(username=nombre_usuario).exists():
            return HttpResponse("El nombre de usuario ya está en uso", status=400)

        # 🔹 3. Validar correo
        if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', correo):
            return HttpResponse("Correo inválido", status=400)
        if Usuarios.objects.filter(correo=correo).exists():
            return HttpResponse("El correo ya está registrado", status=400)

        # 🔹 4. Validar documento
        if not numero_documento.isdigit():
            return HttpResponse("El número de documento debe ser numérico", status=400)
        if len(numero_documento) < 6 or len(numero_documento) > 12:
            return HttpResponse("El número de documento debe tener entre 6 y 12 dígitos", status=400)
        if Usuarios.objects.filter(numerodocumento=numero_documento).exists():
            return HttpResponse("El número de documento ya está registrado", status=400)

        # 🔹 5. Validar contraseñas
        if clave1 != clave2:
            return HttpResponse("Las contraseñas no coinciden", status=400)
        if len(clave1) < 8 or not re.search(r'[A-Z]', clave1) or not re.search(r'[0-9]', clave1):
            return HttpResponse("La contraseña debe tener mínimo 8 caracteres, incluir una mayúscula y un número", status=400)

        # 🔹 6. Validar tipo de documento
        try:
            tipo_documento_obj = TipoDocumentos.objects.get(id=tipo_documento_id)
        except TipoDocumentos.DoesNotExist:
            return HttpResponse("Tipo de documento inválido", status=400)

        # ✅ Crear usuario si todo está correcto
        Usuarios.objects.create(
            username=nombre_usuario,
            nombres=nombres,
            apellidos=apellidos,
            correo=correo,
            idtipodocumento=tipo_documento_obj,
            numerodocumento=numero_documento,
            rol=rol,
            clave=clave1,  # ⚠️ lo ideal sería encriptar con make_password
            estado=estado
        )

        return redirect('registro_exitoso')

    return HttpResponse("Método no permitido", status=405)
def registro_exitoso(request):
    return render(request, 'Logueo/creado.html')
#endregion 
# region Login

def LoginUser(request):
    if request.method == "POST":
        username = request.POST.get("username")
        clave = request.POST.get("clave")
        try:
            usuario = Usuarios.objects.get(username=username, clave=clave)
            if usuario.estado == "Solicitud":
                return render(request, "Logueo/nocreada.html", {
                    "error": "Tu cuenta aún no ha sido aprobada por el administrador.",
                    "tipos_documentos": TipoDocumentos.objects.all()
                })

            # Guardar en sesión
            request.session["usuario_id"] = usuario.id
            request.session["username"] = usuario.username
            request.session["rol"] = usuario.rol

            return redirect("Home")  # o tu página de inicio
        except Usuarios.DoesNotExist:
            return render(request, "Logueo/Logueo.html", {
                "error": "Credenciales inválidas",
                "tipos_documentos": TipoDocumentos.objects.all()
            })
    else:
        # Si es GET y ya está logueado, ir al home
        if request.session.get("usuario_id"):
            return redirect("Home")

        # Si no está logueado, mostrar login
        return redirect("PlantillaLogueo")
    
def nocreada(request):
    return render(request, 'Logueo/nocreado.html')

def logout_view(request):
    request.session.flush()  # Elimina todos los datos de sesión
    return redirect("Home") 
#endregion 

# region ganado
@login_requerido
def TablaGanado(request):
    ganado_list = Ganado.objects.all()
    paginator = Paginator(ganado_list, 9)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    enfermedades = Enfermedades.objects.all()
    vacunas = TablaVacunas.objects.all()
    Razas = TablaRazas.objects.all()
    parcelas = TipoParcela.objects.all()

    return render(request, 'Ganado/Table.html', {
        'page_obj': page_obj,
        "enfermedades": enfermedades,
        "vacunas": vacunas,
        "Razas": Razas,
        'parcelas': parcelas
    })

@login_requerido
def VacasInactivas(request):
    vacas_inactivas = GanadoInactivo.objects.all()
    return render(request, 'Ganado/Inactivas.html', {
        'vacas_inactivas': vacas_inactivas
    })

# Eliminar vacuno (pasar a inactivos)
@login_required
@require_POST
def EliminarVacuno(request, id):
    try:
        vacuno = get_object_or_404(Ganado, id=id)

        # Mover a inactivos
        GanadoInactivo.objects.create(
            codigocria=vacuno.codigocria,
            foto=vacuno.foto,
            crias=vacuno.crias,
            codigoscrias=vacuno.codigoscrias,
            codigopapa=vacuno.codigopapa,
            codigomama=vacuno.codigomama,
            edad=vacuno.edad,
            infovacunas=vacuno.infovacunas,
            enfermedades=vacuno.enfermedades,
            estado=vacuno.estado,
            idparcela=vacuno.idparcela,
            razas=vacuno.razas
        )

        vacuno.delete()
        return JsonResponse({'success': True, 'message': 'Vacuno desactivado correctamente'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# Rehabilitar un vacuno
@login_requerido
def RehabilitarVacuno(request, id):
    vacuno_inactivo = get_object_or_404(GanadoInactivo, id=id)

    # Pasar a la tabla principal
    Ganado.objects.create(
        codigocria=vacuno_inactivo.codigocria,
        foto=vacuno_inactivo.foto,
        crias=vacuno_inactivo.crias,
        codigoscrias=vacuno_inactivo.codigoscrias,
        codigopapa=vacuno_inactivo.codigopapa,
        codigomama=vacuno_inactivo.codigomama,
        edad=vacuno_inactivo.edad,
        infovacunas=vacuno_inactivo.infovacunas,
        enfermedades=vacuno_inactivo.enfermedades,
        estado=vacuno_inactivo.estado,
        idparcela=vacuno_inactivo.idparcela,
        razas=vacuno_inactivo.razas,
        activo=True
    )

    # Eliminar de inactivos
    vacuno_inactivo.delete()
    messages.success(request, "El vacuno fue rehabilitado correctamente.")
    return redirect('VacasInactivas')

@login_requerido
def buscar_codigo_ganado(request):
    q = request.GET.get('q', '').strip()
    resultados = []
    if q:
        resultados = list(Ganado.objects.filter(codigocria__icontains=q).values('id', 'codigocria'))
    # Devuelve id y código
    return JsonResponse([{'id': r['id'], 'codigo': r['codigocria']} for r in resultados], safe=False)


@login_requerido
def ListaRazas(request):
    razas = TablaRazas.objects.all().values('id', 'nombre')
    return JsonResponse(list(razas), safe=False)
@csrf_exempt
@login_requerido
def AgregarRaza(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            razas = TablaRazas(
                nombre=data.get('nombre'),
            )
            razas.save()
            return JsonResponse({'success': True, 'id': razas.id})  # Cambiado a 'success'
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})
#end region

@csrf_exempt
@login_requerido
def registrar_ganado(request):
    if request.method == "POST":
        try:
            codigocria = request.POST.get('CodigoCria')
            foto = request.FILES.get('Foto')
            foto_name = ''
            if foto:
                from django.core.files.storage import default_storage
                foto_name = default_storage.save('ganado/' + foto.name, foto)
            crias = request.POST.get('Crias', '0')
            codigoscrias = request.POST.get('CodigosCrias', '[]')
            codigopapa = request.POST.get('CodigoPapa', '')
            codigomama = request.POST.get('CodigoMama', '')
            edad = request.POST.get('Edad', '')
            infovacunas = request.POST.get('Vacunas', '[]')
            enfermedades = request.POST.get('Enfermedades', '[]')
            estado = request.POST.get('Estado', '')
            idparcela = request.POST.get('IdParcela', '')
            razas = request.POST.get('Razas', '')

            ganado = Ganado.objects.create(
                codigocria=codigocria,
                foto=foto_name,
                crias=crias,
                codigoscrias=codigoscrias,
                codigopapa=codigopapa,
                codigomama=codigomama,
                edad=edad,
                infovacunas=infovacunas,
                enfermedades=enfermedades,
                estado=estado,
                idparcela_id=idparcela,
                razas=razas
            )
            return JsonResponse({'success': True, 'id': ganado.id})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

@csrf_exempt
@login_requerido
def actualizar_ganado(request, id):
    if request.method == "POST":
        try:
            ganado = Ganado.objects.get(id=id)
            ganado.codigocria = request.POST.get('CodigoCria')
            ganado.foto = request.POST.get('Foto', '')
            ganado.crias = request.POST.get('Crias', '0')
            ganado.codigoscrias = request.POST.get('CodigosCrias', '[]')
            ganado.codigopapa = request.POST.get('CodigoPapa', '')
            ganado.codigomama = request.POST.get('CodigoMama', '')
            ganado.edad = request.POST.get('Edad', '')
            ganado.infovacunas = request.POST.get('Vacunas', '[]')
            ganado.enfermedades = request.POST.get('Enfermedades', '[]')
            ganado.codigoscrias = request.POST.get('CodigosCrias', '[]')
            ganado.crias = request.POST.get('Crias', '0')
            ganado.estado = request.POST.get('Estado', '')
            ganado.idparcela_id = request.POST.get('IdParcela', '')
            ganado.razas = request.POST.get('Razas', '')
            ganado.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

@csrf_exempt
@login_requerido
def obtener_ganado(request, id):
    try:
        vacuno = Ganado.objects.get(id=id)
        return JsonResponse({
            'id': vacuno.id,
            'codigocria': vacuno.codigocria,
            'edad': vacuno.edad,
            'estado': vacuno.estado,
            'idparcela': vacuno.idparcela_id,
            'codigomama': vacuno.codigomama,
            'codigopapa': vacuno.codigopapa,
            'razas': vacuno.razas,
            'enfermedades': vacuno.enfermedades,
            'infovacunas': vacuno.infovacunas,
            'codigoscrias': vacuno.codigoscrias,
            'foto': vacuno.foto.url if vacuno.foto else '',
        })
    except Ganado.DoesNotExist:
        return JsonResponse({'error': 'Vacuno no encontrado'}, status=404)

# endregion

# region cultivo
@login_requerido
def TablaCultivo(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        tipo_id = request.POST.get('tipo_id', '').strip()  # Cambiado de tipo_cultivo a tipo_id
        fecha_siembra = request.POST.get('fecha_siembra', '').strip()
        fecha_cosecha = request.POST.get('fecha_cosecha', '').strip()
        cantidad = request.POST.get('cantidad', '').strip()
        foto = request.FILES.get('foto')

        if not nombre or not tipo_id or not fecha_siembra or not fecha_cosecha or not cantidad or not foto:
            return JsonResponse({'error': 'Todos los campos son obligatorios, incluyendo la imagen.'}, status=400)

        cultivo = Cultivo(
            nombre=nombre,
            tipo_id=tipo_id,
            fecha_siembra=fecha_siembra,
            fecha_cosecha=fecha_cosecha,
            cantidad=cantidad,
            foto=foto,
        )
        cultivo.save()
        return JsonResponse({'message': 'Cultivo agregado con éxito'})

    cultivos_list = Cultivo.objects.select_related('tipo').all()
    paginator = Paginator(cultivos_list, 6)  # 6 cultivos por página (puedes ajustar)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'Cultivo/Table.html', {
        'page_obj': page_obj,
        'tipos_cultivo': TipoCultivo.objects.all()  # Pasar los tipos disponibles al template
    })
@login_requerido
def obtener_cultivo(request, id):
    try:
        cultivo = Cultivo.objects.select_related('tipo').get(pk=id)
        return JsonResponse({
            'id': cultivo.id,
            'nombre': cultivo.nombre,
            'tipo_id': cultivo.tipo_id,
            'tipo_nombre': cultivo.tipo.nombre_tipo if cultivo.tipo else '',
            'fecha_siembra': cultivo.fecha_siembra.strftime('%Y-%m-%d'),
            'fecha_cosecha': cultivo.fecha_cosecha.strftime('%Y-%m-%d'),
            'cantidad': cultivo.cantidad,
            'foto': cultivo.foto.url if cultivo.foto else '',
        })
    except Cultivo.DoesNotExist:
        return JsonResponse({'error': 'Cultivo no encontrado'}, status=404)
  
def obtener_fertilizaciones(request, cultivo_id):
    cultivo = Cultivo.objects.get(id=cultivo_id)

    fertilizaciones = Fertilizacion.objects.filter(cultivo_id=cultivo_id).values(
        'fecha', 'fertilizante', 'observaciones', 'tipo', 'dosis'  # ← se agregan
    )

    return JsonResponse({
        'fertilizaciones': list(fertilizaciones),
        'fecha_siembra': cultivo.fecha_siembra.strftime('%Y-%m-%d'),
        'fecha_cosecha': cultivo.fecha_cosecha.strftime('%Y-%m-%d'),
    })


@csrf_exempt
def agregar_fertilizacion(request, cultivo_id):
    if request.method == 'POST':
        fecha = request.POST.get('fecha')
        fertilizante = request.POST.get('fertilizante')
        tipo = request.POST.get('tipo')
        dosis = request.POST.get('dosis')
        observaciones = request.POST.get('observaciones', '')

        # Validaciones básicas
        if not (fecha and fertilizante and tipo and dosis):
            return HttpResponseBadRequest('Faltan campos obligatorios.')

        try:
            cultivo = Cultivo.objects.get(id=cultivo_id)
            fecha_fert = datetime.strptime(fecha, '%Y-%m-%d').date()

            if fecha_fert < cultivo.fecha_siembra:
                return HttpResponseBadRequest('La fecha de fertilización no puede ser anterior a la siembra.')
            if fecha_fert > cultivo.fecha_cosecha:
                return HttpResponseBadRequest('La fecha de fertilización no puede ser posterior a la cosecha.')

            Fertilizacion.objects.create(
                cultivo=cultivo,
                fecha=fecha_fert,
                fertilizante=fertilizante,
                tipo=tipo,
                dosis=dosis,
                observaciones=observaciones
            )

            return JsonResponse({'message': 'Fertilización registrada con éxito'})

        except Cultivo.DoesNotExist:
            return HttpResponseBadRequest('Cultivo no encontrado')
        except Exception as e:
            return HttpResponseBadRequest(str(e))

@login_requerido
def editar_cultivo(request):
    if request.method == 'POST':
        cultivo_id = request.POST.get('id')
        nombre = request.POST.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'error': 'El nombre del cultivo no puede estar vacío'}, status=400)
        try:
            cultivo = Cultivo.objects.get(pk=cultivo_id)
            cultivo.nombre = request.POST.get('nombre')
            cultivo.tipo_id = request.POST.get('tipo_id')  # Cambiado de tipo_cultivo a tipo_id
            cultivo.fecha_siembra = request.POST.get('fecha_siembra')
            cultivo.fecha_cosecha = request.POST.get('fecha_cosecha')
            cultivo.cantidad = request.POST.get('cantidad')
            
            if 'foto' in request.FILES:
                cultivo.foto = request.FILES['foto']
                
            
            cultivo.save()
            return JsonResponse({'message': 'Cultivo editado con éxito'})
        except Cultivo.DoesNotExist:
            return JsonResponse({'error': 'Cultivo no encontrado'}, status=404)

    return JsonResponse({'error': 'Método no permitido'}, status=405)
@login_requerido
def eliminar_cultivo(request):
    cultivo_id = request.POST.get('id')
    try:
        cultivo = Cultivo.objects.get(pk=cultivo_id)
        cultivo.delete()
        return JsonResponse({'message': 'Cultivo eliminado con éxito'})
    except Cultivo.DoesNotExist:
        return JsonResponse({'error': 'El cultivo no existe'}, status=404)

@login_requerido
def InfoBuscador(request,TipoBusqueda,valor):
    if TipoBusqueda == 'documento':
        ganado = Ganado.objects.filter(documento__icontains=valor)
    elif TipoBusqueda == 'nombre':
        ganado = Ganado.objects.filter(nombre__icontains=valor)
    elif TipoBusqueda == 'apellido':
        ganado = Ganado.objects.filter(apellido__icontains=valor)
    elif TipoBusqueda == 'celular':
        ganado = Ganado.objects.filter(celular__icontains=valor)
    elif TipoBusqueda == 'direccion':
        ganado = Ganado.objects.filter(direccion__icontains=valor)
    elif TipoBusqueda == 'cargo':
        ganado = Ganado.objects.filter(cargo__icontains=valor)
    else:
        ganado = Ganado.objects.none()
    
    ganadojson = serialize('json', ganado)
    return HttpResponse(ganadojson, content_type='application/json')
@login_requerido
def obtener_tipoCultivos(request):
    tipos = list(TipoCultivo.objects.values("id", "nombre_tipo"))
    return JsonResponse(tipos, safe=False)

@csrf_exempt
@login_requerido
def agregar_tipoCultivo(request):
    if request.method == "POST":
        data = json.loads(request.body)
        nombre = data.get("nombre_tipo", "").strip()
        if nombre:
            nuevo = TipoCultivo(nombre_tipo=nombre)
            nuevo.save()
            return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)

@csrf_exempt
def eliminar_tipoCultivo(request, id):
    if request.method == "DELETE":
        TipoCultivo.objects.filter(id=id).delete()
        return JsonResponse({"success": True})
    return JsonResponse({"success": False}, status=400)

def obtener_notificaciones(request):
    hoy = date.today()
    notificaciones = []

    cultivos = Cultivo.objects.all()

    for cultivo in cultivos:
        # Notificación por cosecha próxima
        if 0 <= (cultivo.fecha_cosecha - hoy).days <= 3:
            notificaciones.append({
                "tipo": "cosecha",
                "mensaje": f' El cultivo "{cultivo.nombre}" está cerca de su fecha de cosecha ({cultivo.fecha_cosecha})'
            })

        # Notificación por inactividad
        tiene_fertilizaciones = Fertilizacion.objects.filter(cultivo=cultivo).exists()
        if not tiene_fertilizaciones:
            notificaciones.append({
                "tipo": "inactividad",
                "mensaje": f'El cultivo "{cultivo.nombre}" nunca ha sido fertilizado.'
            })

    return JsonResponse({"notificaciones": notificaciones})

# endregion

#RegionParcelas
@login_requerido
def agregar_parcela(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            parcela = TipoParcela(
                nombre=data.get('nombre'),
                estado=data.get('estado')
            )
            parcela.save()
            return JsonResponse({'success': True, 'id': parcela.id})  # Cambiado a 'success'
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Método no permitido'})
@login_requerido
def listar_parcelas(request):
    parcelas = TipoParcela.objects.all().values('id', 'nombre', 'estado')
    return JsonResponse(list(parcelas), safe=False)

@csrf_exempt
@require_POST
@login_requerido
def activar(request, registro_id):
    try:
        parcela = TipoParcela.objects.get(id=registro_id)  # Cambiado a TipoParcela
        # Cambiar el estado (1 a 2 o 2 a 1)
        parcela.estado = 2 if parcela.estado == 1 else 1
        parcela.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Estado de la parcela actualizado correctamente',
            'nuevo_estado': parcela.estado
        })
    except TipoParcela.DoesNotExist:  # Cambiado a TipoParcela
        return JsonResponse({
            'success': False,
            'message': 'Parcela no encontrada'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
@login_requerido    
def Desactivar(request, registro_id):
    try:
        parcela = TipoParcela.objects.get(id=registro_id)  # Cambiado a TipoParcela
        # Cambiar el estado (1 a 2 o 2 a 1)
        parcela.estado = 2 if parcela.estado == 1 else 2
        parcela.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Estado de la parcela actualizado correctamente',
            'nuevo_estado': parcela.estado
        })
    except TipoParcela.DoesNotExist:  # Cambiado a TipoParcela
        return JsonResponse({
            'success': False,
            'message': 'Parcela no encontrada'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)
#end region


# #region excel 

def PlantillaTablas(request):
    razas = TablaRazas.objects.all()
    campos_ganado = [field.name for field in Ganado._meta.fields if field.name not in ['id', 'foto']]
    return render(request, 'Ganado/excel.html', {
        'razas': razas,
        'campos_disponibles': campos_ganado
    })

def ConsultarVacunos(request):
    try:
        # Validar y obtener parámetros
        limit = int(request.GET.get('limit', 10))
        offset = int(request.GET.get('offset', 0))
        fields = request.GET.get('fields', '')
        selected_fields = [f for f in fields.split(',') if f]
        all_fields = [f.name for f in Ganado._meta.fields]
        
        # Validar campos
        valid_fields = [f for f in selected_fields if f in all_fields]
        if not valid_fields:
            return JsonResponse({'success': True, 'vacunos': [], 'total': 0})

        queryset = Ganado.objects.all()

        # Filtro por raza
        filter_raza = request.GET.get('filter_raza')
        if filter_raza:
            queryset = queryset.filter(razas=filter_raza)

        # Orden dinámico
        order_fields = []
        for key, value in request.GET.items():
            if key.startswith('sort_') and value in ['asc', 'desc']:
                field = key[5:]
                if field in valid_fields:
                    order_fields.append(f'-{field}' if value == 'desc' else field)
        if order_fields:
            queryset = queryset.order_by(*order_fields)

        total = queryset.count()
        vacunos = queryset.only(*valid_fields)[offset:offset+limit]

        data = []
        for vacuno in vacunos:
            item = {}
            for field in valid_fields:
                value = getattr(vacuno, field, '')
                # ForeignKey: mostrar id
                if hasattr(value, 'id'):
                    value = value.id
                item[field] = value
            data.append(item)

        return JsonResponse({'success': True, 'vacunos': data, 'total': total})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def ExportarExcel(request):
    try:
        filename = request.GET.get('filename', 'ganado')
        if not filename:
            return HttpResponse("El nombre del archivo es requerido", status=400)

        if re.search(r'[<>:"/\\|?*]', filename):
            return HttpResponse("El nombre del archivo contiene caracteres no válidos", status=400)

        fields = request.GET.get('fields', '')
        selected_fields = [f for f in fields.split(',') if f]
        all_fields = [f.name for f in Ganado._meta.fields]
        valid_fields = [f for f in selected_fields if f in all_fields]
        if not valid_fields:
            return HttpResponse("No hay columnas válidas", status=400)

        queryset = Ganado.objects.all()
        filter_raza = request.GET.get('filter_raza')
        if filter_raza:
            queryset = queryset.filter(razas=filter_raza)

        order_fields = []
        for key, value in request.GET.items():
            if key.startswith('sort_') and value in ['asc', 'desc']:
                field = key[5:]
                if field in valid_fields:
                    order_fields.append(f'-{field}' if value == 'desc' else field)
        if order_fields:
            queryset = queryset.order_by(*order_fields)

        vacunos = queryset.only(*valid_fields)

        wb = Workbook()
        ws = wb.active
        ws.title = "Ganado"

        # ======================
        # ESTILOS
        # ======================
        header_fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")  # Verde
        header_font = Font(bold=True, color="FFFFFF")
        border_style = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        center_alignment = Alignment(horizontal="center", vertical="center")

        # ======================
        # TÍTULO MERGEADO EN FILA 2
        # ======================
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(valid_fields))
        title_cell = ws.cell(row=2, column=1)
        title_cell.value = "Información del Ganado Registrado"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # ======================
        # ENCABEZADOS EN FILA 3
        # ======================
        for col_idx, field in enumerate(valid_fields, 1):
            cell = ws.cell(row=3, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border_style
            cell.alignment = center_alignment

        # ======================
        # DATOS EN FILA 4+
        # ======================
        for row_idx, vacuno in enumerate(vacunos, 4):  # Comienza desde fila 4
            for col_idx, field in enumerate(valid_fields, 1):
                value = getattr(vacuno, field, '')
                if hasattr(value, 'id'):
                    value = value.id
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border_style

        # ======================
        # AJUSTAR ANCHO
        # ======================
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 3

        # ======================
        # EXPORTAR
        # ======================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        return response

    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)

# # def PlantillaGraficas(request):
# #     return render(request, 'Ganado/graficas.html')

# #Cultivo

def PlantillaCultivos(request):
    tipos = TipoCultivo.objects.all()
    campos_cultivo = [field.name for field in Cultivo._meta.fields if field.name not in ['id', 'foto']]
    return render(request, 'Cultivo/excel.html', {
        'tipos': tipos,
        'campos_disponibles': campos_cultivo
    })

def ConsultarCultivos(request):
    try:
        limit = int(request.GET.get('limit', 10))
        offset = int(request.GET.get('offset', 0))
        fields = request.GET.get('fields', '')
        selected_fields = [f for f in fields.split(',') if f]

        # ✅ Incluir campos extra
        model_fields = [f.name for f in Cultivo._meta.fields]
        extra_fields = ['fecha_fertilizacion', 'dosis_fertilizacion']
        valid_fields = [f for f in selected_fields if f in model_fields or f in extra_fields]

        if not valid_fields:
            return JsonResponse({'success': True, 'cultivos': [], 'total': 0})

        queryset = Cultivo.objects.all()

        # ✅ Filtro por tipo
        filter_tipo = request.GET.get('filter_tipo')
        if filter_tipo:
            queryset = queryset.filter(tipo=filter_tipo)

        # ✅ Orden dinámico
        order_fields = []
        for key, value in request.GET.items():
            if key.startswith('sort_') and value in ['asc', 'desc']:
                field = key[5:]
                if field in model_fields:
                    order_fields.append(f'-{field}' if value == 'desc' else field)
        if order_fields:
            queryset = queryset.order_by(*order_fields)

        total = queryset.count()
        cultivos = queryset[offset:offset + limit]

        data = []
        for cultivo in cultivos:
            item = {}
            for field in valid_fields:
                if field == 'fecha_fertilizacion':
                    fert = Fertilizacion.objects.filter(cultivo=cultivo).order_by('fecha').first()
                    item[field] = fert.fecha if fert else ''
                elif field == 'dosis_fertilizacion':
                    fert = Fertilizacion.objects.filter(cultivo=cultivo).order_by('fecha').first()
                    item[field] = fert.dosis if fert else ''
                else:
                    value = getattr(cultivo, field, '')
                    if hasattr(value, 'nombre_tipo'):  # Relación con TipoCultivo
                        value = value.nombre_tipo
                    item[field] = value
            data.append(item)

        return JsonResponse({'success': True, 'cultivos': data, 'total': total})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def ExportarExcelCultivos(request):
    try:
        filename = request.GET.get('filename', 'cultivos')
        if not filename or re.search(r'[<>:"/\\|?*]', filename):
            return HttpResponse("Nombre de archivo inválido", status=400)

        selected_fields = request.GET.get('fields', '').split(',')
        if not selected_fields:
            return HttpResponse("Selecciona columnas", status=400)

        # Validar campos reales y extra
        model_fields = [f.name for f in Cultivo._meta.fields]
        extra_fields = ['fecha_fertilizacion', 'dosis_fertilizacion']
        valid_fields = [f for f in selected_fields if f in model_fields or f in extra_fields]

        queryset = Cultivo.objects.all()

        # Orden y filtros (opcional)
        for key, value in request.GET.items():
            if key.startswith('sort_') and value in ['asc', 'desc']:
                field = key[5:]
                if field in valid_fields:
                    queryset = queryset.order_by(f"{'-' if value == 'desc' else ''}{field}")

        cultivos = queryset

        # Crear libro
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cultivos"

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(valid_fields))
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "Información de los Cultivos Registrados"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4CAF50")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        for col_idx, field in enumerate(valid_fields, 1):
            cell = ws.cell(row=2, column=col_idx, value=field)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        # Datos
        for row_idx, cultivo in enumerate(cultivos, start=3):
            for col_idx, field in enumerate(valid_fields, 1):
                if field in extra_fields:
                    # Buscar primera fertilización
                    fert = Fertilizacion.objects.filter(cultivo=cultivo).order_by('fecha').first()
                    if field == 'fecha_fertilizacion':
                        value = fert.fecha if fert else ''
                    elif field == 'dosis_fertilizacion':
                        value = fert.dosis if fert else ''
                else:
                    value = getattr(cultivo, field, '')
                    if hasattr(value, 'nombre_tipo'):  # relaciones
                        value = value.nombre_tipo
                ws.cell(row=row_idx, column=col_idx, value=value).border = border

        # Ancho automático
        for col in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

        # Guardar
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
        return response

    except Exception as e:
        return HttpResponse(f"Error: {str(e)}", status=500)
    
def ApiTiposCultivo(request):
    tipos = TipoCultivo.objects.all().values('id', 'nombre_tipo')
    return JsonResponse({'tipos': list(tipos)})



# #endregion



# region notificaciones

def SistemaNotficacionesGmail(request):
    pass


# endregion



# region recuperación de contraseña
def olvidar_contra(request):
    if request.session.get("usuario_id"):
        return redirect("Home")
    
    if request.method == "POST":
        correo = request.POST.get("correo", "").strip().lower()
        
        if not correo:
            return render(request, "cambiocontra/olvidar_contra.html", {"error": "El correo es obligatorio"})
        
        try:
            user = Usuarios.objects.get(correo=correo)
            code = str(random.randint(100000, 999999))
            code_hash = hashlib.sha256(code.encode()).hexdigest()
            expires = timezone.now() + timedelta(minutes=15)

            # Limpiar códigos expirados y códigos previos
            codigo.objects.filter(user_id=user.id).delete()
            
            # Crear nuevo código
            codigo_obj = codigo.objects.create(
                user_id=user.id,
                code_hash=code_hash,
                expires_at=expires,
                attempts=0
            )
            
            request.session["reset_email"] = correo
            request.session["reset_user_id"] = user.id
            
            # Enviar email
            try:
                send_reset_email(user.correo, code)
            except Exception as e:
                print(f"Error enviando email: {e}")
                return render(request, "cambiocontra/olvidar_contra.html", {
                    "error": "Error al enviar el código. Intenta nuevamente."
                })

            return redirect("verificar_codigo")
            
        except Usuarios.DoesNotExist:
            # Por seguridad, redirigir igual
            return redirect("verificar_codigo")
        except Exception as e:
            print(f"Error en olvidar_contra: {e}")
            return render(request, "cambiocontra/olvidar_contra.html", {
                "error": "Error interno del sistema. Intenta nuevamente."
            })
    
    return render(request, "cambiocontra/olvidar_contra.html")

def verificar_codigo(request):
    if not request.session.get("reset_email"):
        return redirect("olvidar_contra")
    
    if request.method == "POST":
        code = request.POST.get("code", "").strip()
        email = request.session.get("reset_email")
        user_id = request.session.get("reset_user_id")
        
        if not code or len(code) != 6 or not code.isdigit():
            return render(request, "cambiocontra/verificar_codigo.html", {
                "error": "El código debe tener exactamente 6 dígitos"
            })
        
        try:
            user = Usuarios.objects.get(id=user_id, correo=email)
            reset_obj = codigo.objects.get(user_id=user.id)
            
            if reset_obj.is_expired():
                reset_obj.delete()
                return render(request, "cambiocontra/verificar_codigo.html", {
                    "error": "El código ha expirado. Solicita uno nuevo."
                })
            
            if not reset_obj.check_code(code):
                reset_obj.attempts += 1
                reset_obj.save()
                
                if reset_obj.attempts >= 3:
                    reset_obj.delete()
                    return render(request, "cambiocontra/verificar_codigo.html", {
                        "error": "Demasiados intentos fallidos. Solicita un nuevo código."
                    })
                
                intentos_restantes = 3 - reset_obj.attempts
                return render(request, "cambiocontra/verificar_codigo.html", {
                    "error": f"Código incorrecto. Te quedan {intentos_restantes} intento(s)."
                })
            
            request.session["verified"] = True
            reset_obj.delete()
            return redirect("restablecer_contra")
            
        except Usuarios.DoesNotExist:
            return render(request, "cambiocontra/verificar_codigo.html", {
                "error": "Sesión expirada. Solicita un nuevo código."
            })
        except codigo.DoesNotExist:
            return render(request, "cambiocontra/verificar_codigo.html", {
                "error": "Código no encontrado. Solicita uno nuevo."
            })
        except Exception as e:
            print(f"Error en verificar_codigo: {e}")
            return render(request, "cambiocontra/verificar_codigo.html", {
                "error": "Error interno del sistema."
            })
    
    return render(request, "cambiocontra/verificar_codigo.html")

def restablecer_contra(request):
    if not request.session.get("verified") or not request.session.get("reset_email"):
        return redirect("olvidar_contra")
    
    if request.method == "POST":
        new_password = request.POST.get("new_password", "").strip()
        confirm_password = request.POST.get("confirm_password", "").strip()
        email = request.session.get("reset_email")
        user_id = request.session.get("reset_user_id")
        
        if not new_password:
            return render(request, "cambiocontra/restablecer_contra.html", {
                "error": "La contraseña es obligatoria"
            })
        
        if new_password != confirm_password:
            return render(request, "cambiocontra/restablecer_contra.html", {
                "error": "Las contraseñas no coinciden"
            })
        
        if len(new_password) < 8 or not re.search(r'[A-Z]', new_password) or not re.search(r'[0-9]', new_password):
            return render(request, "cambiocontra/restablecer_contra.html", {
                "error": "La contraseña debe tener mínimo 8 caracteres, incluir una mayúscula y un número"
            })
        
        try:
            user = Usuarios.objects.get(id=user_id, correo=email)
            # ✅ CAMBIO CLAVE: Guardar en texto plano (sin make_password)
            user.clave = new_password  # ← Texto plano directamente
            user.save()
            
            # Limpiar sesión
            request.session.flush()
            
            messages.success(request, "¡Contraseña cambiada exitosamente! Ya puedes iniciar sesión.")
            return redirect("PlantillaLogueo")
            
        except Usuarios.DoesNotExist:
            messages.error(request, "Error al cambiar la contraseña. Intenta nuevamente.")
            return redirect("olvidar_contra")
        except Exception as e:
            print(f"Error en restablecer_contra: {e}")
            return render(request, "cambiocontra/restablecer_contra.html", {
                "error": "Error interno del sistema."
            })
    
    return render(request, "cambiocontra/restablecer_contra.html")

# endregion